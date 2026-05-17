#!/usr/bin/env python3
"""Create and email a weekly BigChange door-to-door engineer timesheet report."""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import math
import os
import re
import smtplib
import sys
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Iterable

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
INCLUDED_GROUP_NAMES = {
    "1. engineer",
    "2. subcontractor",
    "subcontractor",
    "core team - caretaker",
    "core team - electrical",
    "core team - general maintenance",
    "core team - mechanical",
}
PHANTOM_NAME_PARTS = {"cameron north", "kieran", "tom", "winston"}
EXCLUDED_NAME_TOKENS = {"tech", "hk"}
COMPLETION_STATUS_IDS = {12, 13}
START_TRAVEL_STATUS_ID = 8
STARTED_STATUS_ID = 10
OUTPUT_COLUMNS = [
    "Engineer",
    "Date",
    "Day",
    "Start",
    "Finish",
    "Adjusted Hrs",
    "Deduction Applied",
    "Original Time / Deduction Reason",
    "Attention",
    "Attention Details",
    "Jobs",
    "Home Address",
    "First Job Postcode",
    "Distance Miles",
    "Original Start",
    "Original Finish",
    "Start Source",
    "Finish Source",
]


class ReportError(RuntimeError):
    """Raised when the report cannot be produced."""


@dataclass(frozen=True)
class Config:
    bigchange_auth_mode: str
    bigchange_base_url: str
    bigchange_api_key: str
    bigchange_username: str
    bigchange_password: str
    smtp_host: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    smtp_from_email: str
    smtp_from_name: str
    smtp_to_email: str
    jobs_fallback_date_option_id: int | None

    @classmethod
    def from_env(cls) -> "Config":
        def required(name: str) -> str:
            value = os.getenv(name)
            if not value:
                raise ReportError(f"Missing required environment variable: {name}")
            return value

        fallback_raw = os.getenv("BIGCHANGE_JOBS_FALLBACK_DATE_OPTION_ID", "2").strip()
        fallback = int(fallback_raw) if fallback_raw else None
        return cls(
            bigchange_auth_mode=os.getenv("BIGCHANGE_AUTH_MODE", "api_key"),
            bigchange_base_url=required("BIGCHANGE_BASE_URL"),
            bigchange_api_key=required("BIGCHANGE_API_KEY"),
            bigchange_username=required("BIGCHANGE_USERNAME"),
            bigchange_password=required("BIGCHANGE_PASSWORD"),
            smtp_host=required("SMTP_HOST"),
            smtp_port=int(os.getenv("SMTP_PORT", "587")),
            smtp_username=required("SMTP_USERNAME"),
            smtp_password=required("SMTP_PASSWORD"),
            smtp_from_email=required("SMTP_FROM_EMAIL"),
            smtp_from_name=os.getenv("SMTP_FROM_NAME", required("SMTP_FROM_EMAIL")),
            smtp_to_email=required("SMTP_TO_EMAIL"),
            jobs_fallback_date_option_id=fallback,
        )


class BigChangeClient:
    def __init__(self, config: Config) -> None:
        if config.bigchange_auth_mode.lower() != "api_key":
            raise ReportError("Only BIGCHANGE_AUTH_MODE=api_key is supported for legacy Web Services")

        self.config = config
        self.session = requests.Session()
        token = base64.b64encode(
            f"{config.bigchange_username}:{config.bigchange_password}".encode("utf-8")
        ).decode("ascii")
        self.session.headers.update(
            {
                "Accept": "application/json",
                "Authorization": f"Basic {token}",
                "User-Agent": "weekly-door-to-door-timesheet-report/1.0",
            }
        )

    def call(self, action: str, **params: Any) -> Any:
        query = {"action": action, "key": self.config.bigchange_api_key}
        query.update({k: v for k, v in params.items() if v is not None})
        response = self.session.get(self.config.bigchange_base_url, params=query, timeout=60)
        response.raise_for_status()
        payload = response.json()
        code = payload.get("Code")
        result = payload.get("Result")
        if code != 0:
            raise ReportError(f"BigChange {action} failed: {result}")
        if result == "No results" or result is None:
            return []
        return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--today",
        help="Override today's date for report-period calculation (YYYY-MM-DD).",
    )
    parser.add_argument(
        "--output-dir",
        default="reports",
        help="Directory for generated XLSX reports.",
    )
    parser.add_argument(
        "--dry-run-email",
        action="store_true",
        help="Create the workbook but do not send email.",
    )
    return parser.parse_args()


def previous_monday_to_friday(today: dt.date) -> tuple[dt.date, dt.date]:
    days_since_friday = (today.weekday() - 4) % 7
    if days_since_friday == 0:
        days_since_friday = 7
    end = today - dt.timedelta(days=days_since_friday)
    start = end - dt.timedelta(days=4)
    return start, end


def parse_datetime(value: Any) -> dt.datetime | None:
    if not value:
        return None
    if isinstance(value, dt.datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return dt.datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def fmt_time(value: dt.datetime | None) -> str:
    return value.strftime("%H:%M") if value else ""


def fmt_date(value: dt.date) -> str:
    return value.strftime("%Y-%m-%d")


def same_day(value: dt.datetime | None, day: dt.date) -> bool:
    return bool(value and value.date() == day)


def clean_resource_name(label: str) -> str:
    name = re.sub(r"\s+-\s+[^-]+$", "", label or "").strip()
    # BigChange resource labels often include diary/category prefixes such as GM. or E.
    name = re.sub(r"^[A-Z]{1,3}\.\s+", "", name).strip()
    name = re.sub(r"\s+", " ", name)
    return name


def postcode_suffix(label: str) -> str:
    match = re.search(r"\s+-\s+([A-Z]{1,2}\d[A-Z\d]?)\b", label or "", flags=re.I)
    return match.group(1).upper() if match else ""


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (value or "").lower()).strip()


def should_ignore_resource(label: str) -> bool:
    low = (label or "").lower().strip()
    if low.startswith("z."):
        return True
    tokens = set(normalize_name(label).split())
    if any(token.startswith(excluded) for token in tokens for excluded in EXCLUDED_NAME_TOKENS):
        return True
    return any(part in low for part in PHANTOM_NAME_PARTS)


def as_list(result: Any) -> list[dict[str, Any]]:
    if isinstance(result, list):
        return [row for row in result if isinstance(row, dict)]
    if isinstance(result, dict):
        return [result]
    return []


def get_resources(client: BigChangeClient) -> list[dict[str, Any]]:
    groups = as_list(client.call("ResourceGroups"))
    included_group_ids = {
        row.get("id")
        for row in groups
        if str(row.get("label", "")).strip().lower() in INCLUDED_GROUP_NAMES
    }
    if not included_group_ids:
        raise ReportError("Could not find BigChange engineer/subcontractor resource groups")

    resources = []
    for row in as_list(client.call("Resources")):
        label = str(row.get("label") or "")
        if row.get("ResourceGroupId") not in included_group_ids:
            continue
        if should_ignore_resource(label):
            continue
        row["CleanName"] = clean_resource_name(label)
        row["PostcodeSuffix"] = postcode_suffix(label)
        resources.append(row)
    resources.sort(key=lambda r: normalize_name(str(r.get("CleanName") or r.get("label") or "")))
    return resources


def date_range(start: dt.date, end: dt.date) -> Iterable[dt.date]:
    current = start
    while current <= end:
        yield current
        current += dt.timedelta(days=1)


def job_sort_key(job: dict[str, Any]) -> dt.datetime:
    return (
        parse_datetime(job.get("PlannedStart"))
        or parse_datetime(job.get("RealStart"))
        or dt.datetime.max
    )


def fetch_jobs_with_date_option(
    client: BigChangeClient, day: dt.date, date_option_id: int
) -> list[dict[str, Any]]:
    jobs: list[dict[str, Any]] = []
    page = 0
    page_size = 500
    while page < 50:
        result = client.call(
            "JobsList",
            Start=f"{fmt_date(day)} 00:00:00",
            End=f"{fmt_date(day)} 23:59:59",
            IncludeTime="true",
            DateOptionId=date_option_id,
            Allocated=1,
            Unallocated=0,
            ExcludeNullPlannedDates="true",
            Page=page,
            PageSize=page_size,
            IncludeAssistants="true",
        )
        rows = as_list(result)
        jobs.extend(rows)
        if len(rows) < page_size:
            break
        page += 1
    return jobs


def fetch_jobs_for_day(client: BigChangeClient, config: Config, day: dt.date) -> list[dict[str, Any]]:
    requested_jobs = fetch_jobs_with_date_option(client, day, 0)
    if requested_jobs or config.jobs_fallback_date_option_id is None:
        return requested_jobs

    # The legacy API can return no rows for DateOptionId=0 even when planned allocations exist.
    return fetch_jobs_with_date_option(client, day, config.jobs_fallback_date_option_id)


def fetch_journeys_for_day(client: BigChangeClient, day: dt.date) -> list[dict[str, Any]]:
    return as_list(
        client.call(
            "Journeys",
            Start=f"{fmt_date(day)} 00:00:00",
            End=f"{fmt_date(day)} 23:59:59",
        )
    )


def group_by_resource_name(rows: Iterable[dict[str, Any]], field: str = "Resource") -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        name = str(row.get(field) or "").strip()
        if not name:
            continue
        grouped.setdefault(name, []).append(row)
    return grouped


def is_active_job(job: dict[str, Any]) -> bool:
    resource = str(job.get("Resource") or "").strip()
    if not resource:
        return False
    status = str(job.get("Status") or "").strip().lower()
    return "cancel" not in status


def fetch_active_resource_names(
    client: BigChangeClient,
    config: Config,
    active_start: dt.date,
    active_end: dt.date,
) -> set[str]:
    active_names: set[str] = set()
    date_option_id = config.jobs_fallback_date_option_id or 0
    for day in date_range(active_start, active_end):
        for job in fetch_jobs_with_date_option(client, day, date_option_id):
            if not is_active_job(job):
                continue
            resource_name = str(job.get("Resource") or "").strip()
            active_names.add(normalize_name(resource_name))
            active_names.add(normalize_name(clean_resource_name(resource_name)))
    return {name for name in active_names if name}


def best_contact_match(clean_name: str, candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None
    target_tokens = set(normalize_name(clean_name).split())

    def score(candidate: dict[str, Any]) -> tuple[int, int]:
        label = str(candidate.get("label") or "")
        tokens = set(normalize_name(label).split())
        overlap = len(target_tokens & tokens)
        exact = int(normalize_name(label) == normalize_name(clean_name))
        return (exact, overlap)

    return max(candidates, key=score)


def contact_search_terms(clean_name: str) -> list[str]:
    terms = [clean_name]
    parts = clean_name.split()
    if len(parts) >= 2:
        terms.append(" ".join(parts[-2:]))
        terms.append(parts[-1])
    seen: set[str] = set()
    usable = []
    for term in terms:
        term = term.strip()
        if len(term) >= 3 and term.lower() not in seen:
            seen.add(term.lower())
            usable.append(term)
    return usable


def lookup_home(client: BigChangeClient, resource: dict[str, Any]) -> dict[str, Any]:
    label = str(resource.get("label") or "")
    clean_name = str(resource.get("CleanName") or clean_resource_name(label))
    contact_detail: dict[str, Any] | None = None
    lookup_note = ""

    for term in contact_search_terms(clean_name):
        try:
            result = client.call("ContactList", Term=term, Page=0, PageSize=10)
        except ReportError:
            continue
        match = best_contact_match(clean_name, as_list(result))
        contact_id = match.get("id") if match else None
        if contact_id:
            detail = client.call("ContactDetail", ContactId=contact_id)
            if isinstance(detail, dict) and detail.get("PostCode"):
                contact_detail = detail
                break

    if contact_detail:
        address_parts = [
            contact_detail.get("Street"),
            contact_detail.get("Town"),
            contact_detail.get("PostCode"),
            contact_detail.get("Country"),
        ]
        return {
            "postcode": str(contact_detail.get("PostCode") or "").strip(),
            "address": ", ".join(str(part).strip() for part in address_parts if part),
            "lat": contact_detail.get("Lat"),
            "lng": contact_detail.get("Lng"),
            "fallback": False,
            "note": "CRM ContactDetail",
        }

    suffix = str(resource.get("PostcodeSuffix") or "")
    if "mohammed timami" in normalize_name(clean_name) and not suffix:
        suffix = "M8"
    elif "mohammed timami" in normalize_name(clean_name):
        suffix = "M8"

    if suffix:
        lookup_note = f"No CRM address found; using resource postcode suffix/outcode {suffix}"
        return {
            "postcode": suffix,
            "address": lookup_note,
            "lat": None,
            "lng": None,
            "fallback": True,
            "note": lookup_note,
        }

    lookup_note = "No CRM address or resource postcode suffix found"
    return {
        "postcode": "",
        "address": lookup_note,
        "lat": None,
        "lng": None,
        "fallback": True,
        "note": lookup_note,
    }


def fetch_status_history(client: BigChangeClient, job: dict[str, Any]) -> list[dict[str, Any]]:
    job_id = job.get("JobId")
    if not job_id:
        return []
    return as_list(client.call("JobStatusHistory", JobId=job_id))


def status_id(row: dict[str, Any]) -> int | None:
    value = row.get("JobStatusID")
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def status_time(row: dict[str, Any]) -> dt.datetime | None:
    return parse_datetime(row.get("JobStatusDate"))


def find_first_status(
    history: list[dict[str, Any]],
    wanted_status_id: int,
    day: dt.date | None = None,
    before: dt.datetime | None = None,
) -> dt.datetime | None:
    times: list[dt.datetime] = []
    for row in history:
        when = status_time(row)
        if status_id(row) != wanted_status_id or not when:
            continue
        if day and when.date() != day:
            continue
        if before and when > before:
            continue
        times.append(when)
    return min(times) if times else None


def find_latest_completion_same_day(
    histories: dict[int, list[dict[str, Any]]], day: dt.date
) -> dt.datetime | None:
    completions: list[dt.datetime] = []
    for history in histories.values():
        for row in history:
            when = status_time(row)
            if status_id(row) in COMPLETION_STATUS_IDS and same_day(when, day):
                completions.append(when)
    return max(completions) if completions else None


def haversine_miles(lat1: Any, lon1: Any, lat2: Any, lon2: Any) -> float | None:
    try:
        lat1_f, lon1_f, lat2_f, lon2_f = map(float, (lat1, lon1, lat2, lon2))
    except (TypeError, ValueError):
        return None
    radius_miles = 3958.8
    phi1, phi2 = math.radians(lat1_f), math.radians(lat2_f)
    d_phi = math.radians(lat2_f - lat1_f)
    d_lambda = math.radians(lon2_f - lon1_f)
    a = (
        math.sin(d_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    )
    return 2 * radius_miles * math.asin(math.sqrt(a))


def actual_job_start(
    job: dict[str, Any], history: list[dict[str, Any]], day: dt.date
) -> dt.datetime | None:
    started = find_first_status(history, STARTED_STATUS_ID, day)
    if started:
        return started

    real_start = parse_datetime(job.get("RealStart"))
    return real_start if same_day(real_start, day) else None


def first_actual_job(
    jobs: list[dict[str, Any]], histories: dict[int, list[dict[str, Any]]], day: dt.date
) -> tuple[dict[str, Any] | None, dt.datetime | None]:
    for job in jobs:
        job_id = job.get("JobId")
        history = histories.get(int(job_id), []) if job_id else []
        started = actual_job_start(job, history, day)
        if started:
            return job, started
    return None, None


def travel_start_for_job(
    job: dict[str, Any] | None,
    history: list[dict[str, Any]],
    day: dt.date,
    reference_start: dt.datetime | None,
) -> dt.datetime | None:
    if not job or not reference_start:
        return None

    starts: list[dt.datetime] = []
    for row in history:
        when = status_time(row)
        if status_id(row) == START_TRAVEL_STATUS_ID and same_day(when, day) and when < reference_start:
            starts.append(when)
    return min(starts) if starts else None


def distance_to_job(home: dict[str, Any], first_job: dict[str, Any] | None) -> float | None:
    if not first_job:
        return None
    return haversine_miles(
        home.get("lat"),
        home.get("lng"),
        first_job.get("JobContactLatitude"),
        first_job.get("JobContactLongitude"),
    )


def calculate_deduction(
    original_start: dt.datetime | None,
    first_job: dict[str, Any] | None,
    gap_target: dt.datetime | None,
    home: dict[str, Any],
    has_pre_start_evidence: bool,
) -> tuple[int, str, float | None]:
    distance = distance_to_job(home, first_job)
    if not first_job:
        return 0, "No valid travel gap: no jobs allocated", None
    if not original_start:
        return 0, "No valid travel gap: no actual start/travel time found", distance
    if not has_pre_start_evidence:
        return 0, "No valid travel gap: no tracking journey or start travel evidence", distance

    if not gap_target or gap_target <= original_start:
        return 0, "No valid travel gap: first job start/planned time is unavailable or before start", distance

    actual_gap = int((gap_target - original_start).total_seconds() // 60)
    if actual_gap <= 0:
        return 0, "Journey looks tight: no positive travel/pre-start gap", distance

    if distance is None:
        return (
            0,
            f"Distance unavailable: actual travel/pre-start gap {actual_gap} mins",
            None,
        )

    allowance = distance / 35 * 60 + 15
    if actual_gap <= allowance:
        return (
            0,
            "Journey looks tight/no deduction needed: "
            f"actual gap {actual_gap} mins, allowance {allowance:.0f} mins, distance {distance:.1f} miles",
            distance,
        )

    surplus = actual_gap - allowance
    deduction = int(math.floor((surplus / 2) / 15) * 15)
    deduction = min(deduction, 60)
    if deduction <= 0:
        return (
            0,
            "No deduction needed after 15-minute rounding: "
            f"actual gap {actual_gap} mins, allowance {allowance:.0f} mins, distance {distance:.1f} miles",
            distance,
        )

    reason = (
        f"Original start {fmt_time(original_start)}; adjusted start "
        f"{fmt_time(original_start + dt.timedelta(minutes=deduction))}; deduction {deduction} mins; "
        f"actual travel/pre-start gap {actual_gap} mins; reasonable allowance {allowance:.0f} mins; "
        f"approx distance {distance:.1f} miles"
    )
    return deduction, reason, distance


def job_label(job: dict[str, Any]) -> str:
    ref = str(job.get("Ref") or job.get("JobId") or "Job").strip()
    planned = parse_datetime(job.get("PlannedStart"))
    planned_text = planned.strftime("%H:%M") if planned else ""
    postcode = str(job.get("Postcode") or "").strip()
    job_type = str(job.get("Type") or "").strip()
    pieces = [ref]
    if planned_text:
        pieces.append(planned_text)
    if postcode:
        pieces.append(postcode)
    if job_type:
        pieces.append(job_type)
    return " / ".join(pieces)


def attention_for_jobs(
    jobs: list[dict[str, Any]], histories: dict[int, list[dict[str, Any]]]
) -> tuple[str, str]:
    details: list[str] = []
    for job in jobs:
        planned_end = parse_datetime(job.get("PlannedEnd"))
        if not planned_end:
            continue
        job_id = job.get("JobId")
        history = histories.get(int(job_id), []) if job_id else []
        for row in history:
            when = status_time(row)
            if status_id(row) not in COMPLETION_STATUS_IDS or not when:
                continue
            minutes_late = int((when - planned_end).total_seconds() // 60)
            if minutes_late > 120:
                ref = str(job.get("Ref") or job.get("JobId") or "Job")
                details.append(
                    f"{ref}: planned end {planned_end.strftime('%Y-%m-%d %H:%M')}, "
                    f"actual completion {when.strftime('%Y-%m-%d %H:%M')}, {minutes_late} mins late"
                )
    if details:
        return "ATTENTION NEEDED", "; ".join(details)
    return "", ""


def build_day_row(
    client: BigChangeClient,
    resource: dict[str, Any],
    day: dt.date,
    jobs: list[dict[str, Any]],
    journeys: list[dict[str, Any]],
    home: dict[str, Any],
    status_cache: dict[int, list[dict[str, Any]]],
) -> dict[str, Any]:
    clean_name = str(resource.get("CleanName") or resource.get("label") or "")
    if not jobs:
        return {
            "Engineer": clean_name,
            "Date": fmt_date(day),
            "Day": day.strftime("%A"),
            "Start": "",
            "Finish": "",
            "Adjusted Hrs": 0,
            "Deduction Applied": "0 mins",
            "Original Time / Deduction Reason": "No valid travel gap: No jobs allocated",
            "Attention": "",
            "Attention Details": "",
            "Jobs": "No jobs allocated",
            "Home Address": home.get("address", ""),
            "First Job Postcode": "",
            "Distance Miles": "",
            "Original Start": "",
            "Original Finish": "",
            "Start Source": "",
            "Finish Source": "",
            "_attention": False,
            "_deduction_minutes": 0,
        }

    jobs = sorted(jobs, key=job_sort_key)
    histories: dict[int, list[dict[str, Any]]] = {}
    for job in jobs:
        job_id = job.get("JobId")
        if not job_id:
            continue
        job_id_int = int(job_id)
        if job_id_int not in status_cache:
            status_cache[job_id_int] = fetch_status_history(client, job)
        histories[job_id_int] = status_cache[job_id_int]

    first_planned_job = jobs[0]
    first_job_id = first_planned_job.get("JobId")
    first_job_history = histories.get(int(first_job_id), []) if first_job_id else []
    first_started = actual_job_start(first_planned_job, first_job_history, day)
    first_planned_start = parse_datetime(first_planned_job.get("PlannedStart"))
    if not same_day(first_planned_start, day):
        first_planned_start = None
    first_start_reference = first_started or first_planned_start

    journey_starts = [
        parse_datetime(row.get("Start"))
        for row in journeys
        if same_day(parse_datetime(row.get("Start")), day)
    ]
    journey_starts = [value for value in journey_starts if value]
    journey_starts = [
        value for value in journey_starts if first_start_reference and value < first_start_reference
    ]

    original_start: dt.datetime | None = None
    start_source = ""
    if journey_starts:
        original_start = min(journey_starts)
        start_source = "Tracking journey"
    else:
        travel_start = travel_start_for_job(
            first_planned_job, first_job_history, day, first_start_reference
        )
        if travel_start:
            original_start = travel_start
            start_source = "Start travel pressed"
        elif first_started:
            original_start = first_started
            start_source = "First job started"
        elif first_planned_start:
            original_start = first_planned_start
            start_source = "Planned start only"
        else:
            start_source = "No actual start/travel found"

    completion = find_latest_completion_same_day(histories, day)
    if completion:
        original_finish = completion
        finish_source = "Last completion"
    else:
        original_finish = parse_datetime(jobs[-1].get("PlannedEnd"))
        finish_source = "Planned finish only" if original_finish else "No actual finish found"

    deduction_minutes, deduction_reason, distance = calculate_deduction(
        original_start,
        first_planned_job,
        first_start_reference,
        home,
        start_source in {"Tracking journey", "Start travel pressed"},
    )
    adjusted_start = (
        original_start + dt.timedelta(minutes=deduction_minutes)
        if original_start and deduction_minutes
        else original_start
    )
    if deduction_minutes == 0 and original_start and first_planned_job:
        deduction_reason = (
            f"Original start {fmt_time(original_start)}; adjusted start {fmt_time(adjusted_start)}; "
            f"deduction 0 mins; {deduction_reason}"
        )

    adjusted_hours = 0.0
    if adjusted_start and original_finish and original_finish > adjusted_start:
        adjusted_hours = round((original_finish - adjusted_start).total_seconds() / 3600, 2)

    attention, attention_details = attention_for_jobs(jobs, histories)
    first_postcode = str(first_planned_job.get("Postcode") or "").strip()
    return {
        "Engineer": clean_name,
        "Date": fmt_date(day),
        "Day": day.strftime("%A"),
        "Start": fmt_time(adjusted_start),
        "Finish": fmt_time(original_finish),
        "Adjusted Hrs": adjusted_hours,
        "Deduction Applied": f"{deduction_minutes} mins",
        "Original Time / Deduction Reason": deduction_reason,
        "Attention": attention,
        "Attention Details": attention_details,
        "Jobs": "\n".join(job_label(job) for job in jobs),
        "Home Address": home.get("address", ""),
        "First Job Postcode": first_postcode,
        "Distance Miles": round(distance, 1) if distance is not None else "",
        "Original Start": fmt_time(original_start),
        "Original Finish": fmt_time(original_finish),
        "Start Source": start_source,
        "Finish Source": finish_source,
        "_attention": bool(attention),
        "_deduction_minutes": deduction_minutes,
    }


def write_workbook(
    rows: list[dict[str, Any]],
    week_start: dt.date,
    week_end: dt.date,
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    weekly_title = f"{fmt_date(week_start)} to {week_end.strftime('%m-%d')}"
    weekly = workbook.active
    weekly.title = weekly_title[:31]
    weekly.append(OUTPUT_COLUMNS)
    for row in rows:
        weekly.append([row.get(column, "") for column in OUTPUT_COLUMNS])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    attention_fill = PatternFill("solid", fgColor="FFC7CE")
    for sheet in [weekly]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for idx, row in enumerate(rows, start=2):
            if row.get("_attention"):
                sheet.cell(idx, OUTPUT_COLUMNS.index("Finish") + 1).fill = attention_fill
            for cell in sheet[idx]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 24,
        "B": 12,
        "C": 11,
        "D": 10,
        "E": 10,
        "F": 12,
        "G": 18,
        "H": 54,
        "I": 18,
        "J": 54,
        "K": 45,
        "L": 38,
        "M": 18,
        "N": 15,
        "O": 14,
        "P": 14,
        "Q": 20,
        "R": 18,
    }
    for col, width in widths.items():
        weekly.column_dimensions[col].width = width

    summary = workbook.create_sheet("Summary")
    summary.append(["Engineer", "Total Deduction Minutes", "Total Deduction Hours"])
    totals: dict[str, int] = {}
    for row in rows:
        totals[row["Engineer"]] = totals.get(row["Engineer"], 0) + int(row.get("_deduction_minutes", 0))
    for engineer in sorted(totals, key=normalize_name):
        minutes = totals[engineer]
        summary.append([engineer, minutes, round(minutes / 60, 2)])

    attention = workbook.create_sheet("Attention Needed")
    attention.append(["Date", "Engineer", "Finish", "Attention Details", "Jobs"])
    for row in rows:
        if row.get("_attention"):
            attention.append(
                [
                    row.get("Date", ""),
                    row.get("Engineer", ""),
                    row.get("Finish", ""),
                    row.get("Attention Details", ""),
                    row.get("Jobs", ""),
                ]
            )

    for sheet in [summary, attention]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(idx)].width = 28 if idx != 4 else 70

    path = output_dir / f"door_to_door_timesheets_{fmt_date(week_start)}_to_{fmt_date(week_end)}.xlsx"
    workbook.save(path)
    return path


def email_body(
    rows: list[dict[str, Any]], week_start: dt.date, week_end: dt.date
) -> str:
    totals: dict[str, int] = {}
    attention_rows: list[dict[str, Any]] = []
    for row in rows:
        totals[row["Engineer"]] = totals.get(row["Engineer"], 0) + int(row.get("_deduction_minutes", 0))
        if row.get("_attention"):
            attention_rows.append(row)

    lines = [
        f"Weekly door-to-door timesheets report: {fmt_date(week_start)} to {fmt_date(week_end)}",
        "",
        "Deductions were applied automatically to the displayed Start time where the travel/pre-start gap exceeded the distance-based allowance.",
        "Red Finish cells need checking where completion was recorded more than 2 hours after planned finish.",
        "",
        "Total deductions by engineer:",
    ]
    for engineer in sorted(totals, key=normalize_name):
        minutes = totals[engineer]
        lines.append(f"- {engineer}: {minutes} mins ({minutes / 60:.2f} hrs)")

    lines.extend(["", "Attention-needed rows:"])
    if attention_rows:
        for row in attention_rows:
            lines.append(
                f"- {row['Date']} {row['Engineer']} finish {row['Finish']}: {row['Attention Details']}"
            )
    else:
        lines.append("- None")
    return "\n".join(lines)


def send_email(
    config: Config,
    workbook_path: Path,
    rows: list[dict[str, Any]],
    week_start: dt.date,
    week_end: dt.date,
) -> None:
    message = EmailMessage()
    message["From"] = f"{config.smtp_from_name} <{config.smtp_from_email}>"
    message["To"] = config.smtp_to_email
    message["Subject"] = (
        f"Nirvana Weekly door-to-door timesheets report - {fmt_date(week_start)} to {fmt_date(week_end)}"
    )
    message.set_content(email_body(rows, week_start, week_end))
    data = workbook_path.read_bytes()
    message.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=workbook_path.name,
    )

    with smtplib.SMTP(config.smtp_host, config.smtp_port, timeout=60) as smtp:
        smtp.starttls()
        smtp.login(config.smtp_username, config.smtp_password)
        smtp.send_message(message)


def build_report(
    client: BigChangeClient,
    config: Config,
    week_start: dt.date,
    week_end: dt.date,
    active_window_end: dt.date,
) -> list[dict[str, Any]]:
    resources = get_resources(client)
    active_window_start = active_window_end - dt.timedelta(days=29)
    active_resource_names = fetch_active_resource_names(
        client, config, active_window_start, active_window_end
    )
    resources = [
        resource
        for resource in resources
        if normalize_name(str(resource.get("label") or "")) in active_resource_names
        or normalize_name(str(resource.get("CleanName") or "")) in active_resource_names
    ]
    homes = {int(resource["id"]): lookup_home(client, resource) for resource in resources}
    rows: list[dict[str, Any]] = []
    status_cache: dict[int, list[dict[str, Any]]] = {}

    jobs_by_day: dict[dt.date, dict[str, list[dict[str, Any]]]] = {}
    journeys_by_day: dict[dt.date, dict[str, list[dict[str, Any]]]] = {}
    for day in date_range(week_start, week_end):
        jobs_by_day[day] = group_by_resource_name(fetch_jobs_for_day(client, config, day))
        journeys_by_day[day] = group_by_resource_name(fetch_journeys_for_day(client, day))

    for resource in resources:
        label = str(resource.get("label") or "")
        clean = str(resource.get("CleanName") or clean_resource_name(label))
        home = homes[int(resource["id"])]
        for day in date_range(week_start, week_end):
            day_jobs = jobs_by_day[day].get(label) or jobs_by_day[day].get(clean) or []
            day_journeys = journeys_by_day[day].get(label) or journeys_by_day[day].get(clean) or []
            rows.append(
                build_day_row(client, resource, day, day_jobs, day_journeys, home, status_cache)
            )
    return rows


def main() -> int:
    args = parse_args()
    config = Config.from_env()
    today = dt.date.fromisoformat(args.today) if args.today else dt.date.today()
    week_start, week_end = previous_monday_to_friday(today)
    client = BigChangeClient(config)

    rows = build_report(client, config, week_start, week_end, today)
    output_path = write_workbook(rows, week_start, week_end, Path(args.output_dir))
    attention_count = sum(1 for row in rows if row.get("_attention"))
    engineers = sorted({row["Engineer"] for row in rows}, key=normalize_name)

    email_status = "skipped"
    if not args.dry_run_email:
        send_email(config, output_path, rows, week_start, week_end)
        email_status = "sent"

    print(f"Report period: {fmt_date(week_start)} to {fmt_date(week_end)}")
    print(f"Engineers included: {len(engineers)}")
    print(f"Rows created: {len(rows)}")
    print(f"Attention flags: {attention_count}")
    print(f"Email: {email_status}")
    print(f"Workbook: {output_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Report failed: {exc}", file=sys.stderr)
        raise SystemExit(1)
